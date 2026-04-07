from flask import Flask, render_template, request, jsonify
from werkzeug.utils import secure_filename
import openpyxl
import os
import json as _json
from datetime import datetime, timedelta
from collections import defaultdict
import traceback

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = '/tmp'

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

DF_META_BY_FLEET = {
    'moto': 72.0,
    'esc': 90.50
}

SYSTEM_NORMALIZE = {
    'hidraulico': 'Hidráulico', 'hidráulico': 'Hidráulico', 'Hidraulico': 'Hidráulico', 'Hidraúlico': 'Hidráulico', 'hidraúlico': 'Hidráulico',
    'eletroeletrônico': 'Eletroeletrônico', 'eletroeletronico': 'Eletroeletrônico', 'Eletroeletronico': 'Eletroeletrônico', 'Eletroeletrônico': 'Eletroeletrônico',
    'implemento': 'Implemento', 'implmento': 'Implemento', 'Implemento': 'Implemento',
    'transmissão': 'Transmissão', 'Transmissão': 'Transmissão',
    'locomoção': 'Locomoção', 'Locomoção': 'Locomoção',
    'direção': 'Direção', 'Direção': 'Direção',
    'lubrificação': 'Lubrificação', 'Lubrificação': 'Lubrificação',
    'ar condicionado': 'Ar Condicionado', 'Ar condicionado': 'Ar Condicionado',
    'motriz': 'Motriz', 'Motriz': 'Motriz',
    'combustivél': 'Combustível', 'combustivel': 'Combustível',
    'sci': 'SCI', 'SCI': 'SCI',
    'slc': 'SLC', 'SLC': 'SLC',
    'giro': 'Giro', 'Giro': 'Giro',
    'estrutura': 'Estrutura', 'Estrutura': 'Estrutura',
    'freio': 'Freio', 'Freio': 'Freio',
}


def parse_duration(dur_val):
    """Parse duration from Excel - handles time, datetime, timedelta, string"""
    if dur_val is None:
        return 0
    try:
        import datetime as dt_mod
        if isinstance(dur_val, timedelta):
            return int(dur_val.total_seconds())
        if isinstance(dur_val, dt_mod.time):
            return dur_val.hour * 3600 + dur_val.minute * 60 + dur_val.second
        if isinstance(dur_val, datetime):
            base = datetime(1900, 1, 1)
            delta = dur_val - base
            return max(0, int(delta.total_seconds()))
        s = str(dur_val).strip()
        if 'day' in s:
            parts = s.split(',')
            days = int(parts[0].split()[0])
            time_parts = parts[1].strip().split(':')
            return days * 86400 + int(time_parts[0]) * 3600 + int(time_parts[1]) * 60 + int(time_parts[2])
        if ':' in s:
            parts = s.split(':')
            if len(parts) >= 2:
                return int(parts[0]) * 3600 + int(parts[1]) * 60 + (int(float(parts[2])) if len(parts) > 2 else 0)
    except:
        pass
    return 0


def parse_time(time_str):
    """Parse time string HH:MM or HH:MM:SS to seconds"""
    if not time_str:
        return None
    try:
        import datetime as dt_mod
        if isinstance(time_str, dt_mod.time):
            return time_str.hour * 3600 + time_str.minute * 60 + time_str.second
        time_str = str(time_str).strip()
        if ':' in time_str:
            parts = time_str.split(':')
            if len(parts) >= 2:
                return int(parts[0]) * 3600 + int(parts[1]) * 60 + (int(float(parts[2])) if len(parts) > 2 else 0)
    except:
        pass
    return 0


def fmt_time(seconds):
    """Format seconds to HH:MM"""
    if not seconds:
        return "0h 0m"
    h = seconds // 3600
    m = (seconds % 3600) // 60
    return '{}h {}m'.format(int(h), int(m))


def to_str(val):
    """Safe string conversion"""
    if val is None:
        return ''
    return str(val).strip()


def to_date(val):
    """Parse date from Excel, reject invalid dates (1900, etc)"""
    if isinstance(val, datetime):
        if val.year < 2000:
            return None
        return val.date()
    try:
        d = datetime.strptime(str(val), '%d/%m/%Y').date()
        if d.year < 2000:
            return None
        return d
    except:
        return None


def normalize_system(system_name):
    """Normalize system name using mapping"""
    if not system_name:
        return ''
    s = to_str(system_name).lower()
    return SYSTEM_NORMALIZE.get(s, to_str(system_name))


def process_fleet(ws, col_offset, fleet_prefix, equipment_ids):
    """
    Process a fleet failure sheet.
    col_offset=0: Moto (A=Data, B=Equipamentos, ...)
    col_offset=1: Escavadeira (B=Data, C=Equipamentos, ...)
    """
    failures = []
    eq_failures = defaultdict(int)
    sys_failures = defaultdict(int)
    subsys_failures = defaultdict(int)
    eq_hours = defaultdict(int)

    header_row = 1
    data_start_row = 2
    max_row = ws.max_row

    for row_idx in range(data_start_row, max_row + 1):
        try:
            col_a = col_offset + 1
            col_b = col_offset + 2
            col_c = col_offset + 3
            col_d = col_offset + 4
            col_e = col_offset + 5
            col_f = col_offset + 6
            col_g = col_offset + 7
            col_h = col_offset + 8

            data_val = ws.cell(row_idx, col_a).value
            eq_val = ws.cell(row_idx, col_b).value
            descr = to_str(ws.cell(row_idx, col_c).value)
            system = to_str(ws.cell(row_idx, col_d).value)
            subsystem = to_str(ws.cell(row_idx, col_e).value)
            hora_inicio = parse_time(ws.cell(row_idx, col_f).value)
            hora_fim = parse_time(ws.cell(row_idx, col_g).value)
            duracao_val = ws.cell(row_idx, col_h).value

            if not data_val or not eq_val:
                continue

            eq_id = to_str(eq_val)
            if eq_id not in equipment_ids:
                continue

            fail_date = to_date(data_val)
            if not fail_date:
                continue

            duracao = parse_duration(duracao_val)
            if duracao <= 0 and hora_fim and hora_inicio:
                duracao = hora_fim - hora_inicio
            if duracao < 0:
                duracao = 0

            system_norm = normalize_system(system)

            failures.append({
                'data': fail_date,
                'eq_id': eq_id,
                'eq_display': fleet_prefix + eq_id,
                'descr': descr,
                'system': system_norm,
                'subsystem': subsystem,
                'duracao': duracao,
            })

            eq_failures[eq_id] += 1
            eq_hours[eq_id] += duracao
            if system_norm:
                sys_failures[system_norm] += 1
            if subsystem:
                subsys_failures[subsystem] += 1
        except Exception as e:
            continue

    return {
        'failures': failures,
        'eq_failures': dict(eq_failures),
        'sys_failures': dict(sys_failures),
        'subsys_failures': dict(subsys_failures),
        'eq_hours': dict(eq_hours),
    }


def parse_df_value(val):
    """Parse DF value - handles decimals (0-1), percentages (>1), text, None"""
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip().lower()
        if s in ('desprogramada', '', '-', 'n/a'):
            return None
        try:
            v = float(s)
        except:
            return None
    else:
        try:
            v = float(val)
        except:
            return None
    if v > 1.5:
        return min(v, 100.0)
    else:
        return v * 100.0


def process_df_sheet(ws, equipment_ids):
    """
    Process DF sheet: daily data with equipment averages and fleet daily DF.
    Handles mixed decimal/percentage values, Desprogramada text, Acumulado rows.
    Returns: daily_df, equipment_averages, fleet_daily_df, acumulado
    """
    daily_df = {}
    equipment_averages = {}
    fleet_daily_df = {}
    acumulado = None

    try:
        header_row = 3
        col_a_val = ws.cell(header_row, 1).value
        if to_str(col_a_val).upper() != 'DATA':
            return daily_df, equipment_averages, fleet_daily_df, acumulado

        eq_cols = []
        for col_idx in range(2, 12):
            eq_id = to_str(ws.cell(header_row, col_idx).value)
            if eq_id and eq_id in equipment_ids:
                eq_cols.append((col_idx, eq_id))

        if not eq_cols:
            return daily_df, equipment_averages, fleet_daily_df, acumulado

        fleet_df_col = None
        if eq_cols:
            last_eq_col = max(col for col, _ in eq_cols)
            fleet_df_col = last_eq_col + 1

        eq_values = defaultdict(list)

        for row_idx in range(4, ws.max_row + 1):
            date_val = ws.cell(row_idx, 1).value
            if not date_val:
                continue

            date_str = to_str(date_val).lower()
            is_acumulado_row = 'acumulado' in date_str

            if is_acumulado_row:
                if fleet_df_col and acumulado is None:
                    fleet_df_val = ws.cell(row_idx, fleet_df_col).value
                    pct = parse_df_value(fleet_df_val)
                    if pct is not None:
                        acumulado = pct
                continue

            if 'total' in date_str or 'media' in date_str:
                continue

            try:
                if isinstance(date_val, datetime):
                    if date_val.year < 2000:
                        continue
                    d = date_val.date()
                else:
                    d = datetime.strptime(str(date_val), '%d/%m/%Y').date()
                date_key = d.strftime('%Y-%m-%d')
            except:
                continue

            daily_record = {}
            valid_count = 0
            for col_idx, eq_id in eq_cols:
                val = ws.cell(row_idx, col_idx).value
                pct = parse_df_value(val)
                if pct is not None:
                    daily_record[eq_id] = pct
                    eq_values[eq_id].append(pct)
                    valid_count += 1

            if valid_count > 0:
                daily_df[date_key] = daily_record

            if fleet_df_col:
                fleet_df_val = ws.cell(row_idx, fleet_df_col).value
                pct = parse_df_value(fleet_df_val)
                if pct is not None:
                    fleet_daily_df[date_key] = pct

        for eq_id in equipment_ids:
            if eq_id in eq_values and eq_values[eq_id]:
                avg = sum(eq_values[eq_id]) / len(eq_values[eq_id])
                equipment_averages[eq_id] = round(avg, 2)
            else:
                equipment_averages[eq_id] = 0
    except Exception as e:
        pass

    return daily_df, equipment_averages, fleet_daily_df, acumulado


def process_excel(filepath):
    """Main processing: reads both fleets + both DF sheets"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    moto_eq = ['7401', '8201', '8202', '8301', '8302', '8303']
    esc_eq = ['9401', '9402', '9403', '9404', '9405', '9406', '9407']

    moto_data = process_fleet(wb['Motoniveladora'], 0, 'MM', moto_eq)
    esc_data = process_fleet(wb['Escavadeira'], 1, 'EM', esc_eq)

    moto_daily_df, moto_eq_avg, moto_fleet_df, moto_acum = process_df_sheet(wb['DF MOTO'], moto_eq)
    esc_daily_df, esc_eq_avg, esc_fleet_df, esc_acum = process_df_sheet(wb['DF ESCAVADEIRA'], esc_eq)

    moto_data['daily_df'] = moto_daily_df
    moto_data['eq_avg'] = moto_eq_avg
    moto_data['fleet_daily_df'] = moto_fleet_df
    moto_data['acumulado'] = moto_acum

    esc_data['daily_df'] = esc_daily_df
    esc_data['eq_avg'] = esc_eq_avg
    esc_data['fleet_daily_df'] = esc_fleet_df
    esc_data['acumulado'] = esc_acum

    return {
        'moto': moto_data,
        'esc': esc_data,
    }


def get_top_failures(sys_failures, count=5):
    """Get top N failures by frequency"""
    sorted_sys = sorted(sys_failures.items(), key=lambda x: x[1], reverse=True)
    total = sum(sys_failures.values())
    top = sorted_sys[:count]
    top_pct = [(name, freq, round(100 * freq / total, 1)) for name, freq in top]
    return top_pct


def get_critical_failures(failures, eq_hours, top_n=3):
    """Extract top critical failure patterns by duration"""
    pattern_data = defaultdict(lambda: {'count': 0, 'total_hours': 0, 'dates': []})

    for f in failures:
        key = f['system'] + ' - ' + f['subsystem'] if f['subsystem'] else f['system']
        if not key.strip().startswith(' - '):
            pattern_data[key]['count'] += 1
            pattern_data[key]['total_hours'] += f['duracao'] / 3600.0
            pattern_data[key]['dates'].append(f['data'].strftime('%d/%m'))

    patterns = [(k, v['count'], v['total_hours'], v['dates'])
                for k, v in pattern_data.items() if v['count'] > 0]
    patterns.sort(key=lambda x: x[2], reverse=True)

    return patterns[:top_n]


def get_pareto_data(sys_failures, top_n=8):
    """Prepare Pareto data for chart: top N + Others"""
    sorted_sys = sorted(sys_failures.items(), key=lambda x: x[1], reverse=True)

    if len(sorted_sys) <= top_n:
        return sorted_sys, []

    top = sorted_sys[:top_n]
    others_count = sum(freq for _, freq in sorted_sys[top_n:])
    others = ('Outros', others_count)

    return top + [others], []


def generate_report(data, custom_meta=None):
    """Generate complete HTML dashboard with fleet selector"""
    meta = custom_meta if custom_meta else DF_META_BY_FLEET
    p = []

    p.append('<!DOCTYPE html>')
    p.append('<html lang="pt-BR">')
    p.append('<head>')
    p.append('<meta charset="UTF-8">')
    p.append('<meta name="viewport" content="width=device-width, initial-scale=1.0">')
    p.append('<title>Relatório de Confiabilidade - Frota Manutenção</title>')
    p.append('<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>')
    p.append('<script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script>')
    p.append('<style>')

    css_rules = [
        '* { margin: 0; padding: 0; box-sizing: border-box; }',
        'body { font-family: Segoe UI, Tahoma, Geneva, Verdana, sans-serif; background: #f5f5f5; color: #333; }',
        '.header { background: linear-gradient(135deg, #1a3a52 0%, #2d5a7a 100%); color: white; padding: 25px 40px; box-shadow: 0 4px 8px rgba(0,0,0,0.1); }',
        '.header h1 { font-size: 28px; margin-bottom: 5px; }',
        '.header .subtitle { font-size: 14px; opacity: 0.9; margin-bottom: 15px; }',
        '.header .meta-info { font-size: 13px; opacity: 0.85; }',
        '.header .fleet-selector { margin-top: 15px; max-width: 400px; }',
        '.header .fleet-selector label { display: block; font-size: 13px; margin-bottom: 6px; }',
        '.header .fleet-selector select { width: 100%; padding: 8px 12px; border-radius: 4px; border: none; font-size: 14px; }',
        '.container { max-width: 1600px; margin: 0 auto; padding: 20px; }',
        '.fleet-content { display: none; }',
        '.fleet-content.active { display: block; }',
        '.kpi-row { display: grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 20px; margin-bottom: 30px; }',
        '.kpi-card { background: white; padding: 20px; border-radius: 6px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #1a3a52; }',
        '.kpi-label { font-size: 11px; color: #888; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 10px; font-weight: 600; }',
        '.kpi-value { font-size: 32px; font-weight: bold; color: #1a3a52; margin-bottom: 8px; }',
        '.kpi-unit { font-size: 14px; color: #aaa; margin-left: 6px; }',
        '.kpi-status { font-size: 12px; padding: 6px 10px; border-radius: 3px; display: inline-block; font-weight: 600; }',
        '.kpi-status.above { background: #d4edda; color: #155724; }',
        '.kpi-status.below { background: #f8d7da; color: #721c24; }',
        '.kpi-status.excellent { background: #d4edda; color: #155724; }',
        '.kpi-status.good { background: #d1ecf1; color: #0c5460; }',
        '.kpi-status.attention { background: #fff3cd; color: #856404; }',
        '.kpi-status.critical { background: #f8d7da; color: #721c24; }',
        '.status-comment { font-size: 13px; color: #666; margin-top: 15px; padding: 12px; background: #f9f9f9; border-left: 3px solid #ddd; border-radius: 3px; }',
        '.section { background: white; padding: 25px; border-radius: 6px; margin-bottom: 25px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }',
        '.section h2 { font-size: 18px; margin-bottom: 20px; color: #1a3a52; border-bottom: 2px solid #1a3a52; padding-bottom: 10px; }',
        '.subsection { margin-bottom: 25px; }',
        '.subsection h3 { font-size: 14px; color: #2d5a7a; margin-bottom: 12px; font-weight: 600; }',
        '.metric-item { display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #f0f0f0; }',
        '.metric-label { color: #666; }',
        '.metric-value { font-weight: 600; color: #1a3a52; }',
        '.chart-container { position: relative; height: 400px; margin-bottom: 20px; }',
        '.eq-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 15px; }',
        '.eq-card { border: 2px solid #ddd; border-left: 5px solid #888; padding: 15px; border-radius: 4px; }',
        '.eq-card.green { border-left-color: #27ae60; background: rgba(39, 174, 96, 0.02); }',
        '.eq-card.orange { border-left-color: #f39c12; background: rgba(243, 156, 18, 0.02); }',
        '.eq-card.red { border-left-color: #e74c3c; background: rgba(231, 76, 60, 0.02); }',
        '.eq-card.blue { border-left-color: #3498db; background: rgba(52, 152, 219, 0.02); }',
        '.eq-name { font-weight: bold; font-size: 16px; margin-bottom: 10px; }',
        '.eq-name.green { color: #27ae60; }',
        '.eq-name.orange { color: #f39c12; }',
        '.eq-name.red { color: #e74c3c; }',
        '.eq-name.blue { color: #3498db; }',
        '.eq-stat { display: flex; justify-content: space-between; font-size: 13px; margin: 6px 0; }',
        '.eq-stat-label { color: #666; }',
        '.eq-stat-value { font-weight: 600; }',
        '.progress-bar { width: 100%; height: 6px; background: #eee; border-radius: 3px; margin-top: 10px; overflow: hidden; }',
        '.progress-fill { height: 100%; background: #1a3a52; transition: width 0.3s; }',
        '.critical-card { border-left: 4px solid #e74c3c; background: rgba(231, 76, 60, 0.02); padding: 15px; margin-bottom: 12px; border-radius: 4px; }',
        '.critical-card h3 { color: #e74c3c; font-size: 14px; margin-bottom: 8px; }',
        '.critical-card p { font-size: 13px; color: #555; margin: 4px 0; }',
        '.days-critical { display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 12px; margin-bottom: 20px; }',
        '.day-card { border: 2px solid #e74c3c; background: rgba(231, 76, 60, 0.05); padding: 12px; border-radius: 4px; }',
        '.day-card-date { font-weight: bold; color: #e74c3c; font-size: 14px; margin-bottom: 6px; }',
        '.day-card-df { font-size: 12px; color: #666; }',
        '.day-card-failure { font-size: 12px; color: #555; margin-top: 6px; padding-top: 6px; border-top: 1px solid #f0f0f0; }',
        '.footer { background: #f0f0f0; padding: 20px; text-align: center; color: #666; font-size: 12px; margin-top: 40px; border-top: 1px solid #ddd; }',
        '.action-badge { display: inline-block; padding: 4px 12px; border-radius: 12px; font-size: 11px; font-weight: bold; margin: 2px; }',
        '.action-critica { background: #e74c3c; color: white; }',
        '.action-alta { background: #f39c12; color: white; }',
        '.action-media { background: #3498db; color: white; }',
        '.pdf-btn { background: #1a3a52; color: white; padding: 12px 24px; border: none; border-radius: 4px; cursor: pointer; font-size: 14px; font-weight: 500; margin-bottom: 20px; transition: background 0.3s; }',
        '.pdf-btn:hover { background: #2d5a7a; }',
    ]

    for rule in css_rules:
        p.append(rule)

    p.append('</style>')
    p.append('</head>')
    p.append('<body>')

    p.append('<div class="header">')
    p.append('<h1>Relatório de Confiabilidade</h1>')
    p.append('<div class="subtitle">Março 2026 | Frota de <span id="header-fleet">Motoniveladoras</span> GMO</div>')
    moto_meta_display = '{:.2f}'.format(meta['moto']).replace('.', ',')
    p.append('<div class="meta-info">Meta de DF: <span id="meta-df">{}%</span></div>'.format(moto_meta_display))
    p.append('<div class="fleet-selector">')
    p.append('<label for="fleet-select">Selecione a Frota:</label>')
    p.append('<select id="fleet-select" onchange="changeFleet(this.value)">')
    p.append('<option value="moto">Motoniveladoras</option>')
    p.append('<option value="esc">Escavadeiras</option>')
    p.append('</select>')
    p.append('</div>')
    p.append('</div>')

    p.append('<div class="container">')

    p.append('<div style="text-align: right; margin-bottom: 20px;">')
    p.append('<button class="pdf-btn" onclick="downloadPDF()">Baixar PDF</button>')
    p.append('</div>')

    p.append(generate_fleet_tab(data['moto'], 'moto', 'Motoniveladoras', meta['moto']))
    p.append(generate_fleet_tab(data['esc'], 'esc', 'Escavadeiras', meta['esc']))

    p.append('</div>')

    p.append('<div class="footer">')
    p.append('Engenharia de Manutenção e Confiabilidade - Setor GRD<br>')
    p.append('Relatório automático gerado em ' + datetime.now().strftime('%d/%m/%Y às %H:%M:%S'))
    p.append('</div>')

    p.append('<script>')
    p.append('function changeFleet(fleetId) {')
    p.append('  document.querySelectorAll(".fleet-content").forEach(el => el.classList.remove("active"));')
    p.append('  document.getElementById("fleet_" + fleetId).classList.add("active");')
    p.append('  var names = {"moto": "Motoniveladoras", "esc": "Escavadeiras"};')
    moto_meta_str = '{:.2f}'.format(meta['moto']).replace('.', ',')
    esc_meta_str = '{:.2f}'.format(meta['esc']).replace('.', ',')
    p.append('  var metas = {{"moto": "{}%", "esc": "{}%"}};'.format(moto_meta_str, esc_meta_str))
    p.append('  document.getElementById("header-fleet").textContent = names[fleetId] || fleetId;')
    p.append('  document.getElementById("meta-df").textContent = metas[fleetId] || "90,50%";')
    p.append('}')
    p.append('function downloadPDF() {')
    p.append('  var active = document.querySelector(".fleet-content.active");')
    p.append('  if (!active) { alert("Selecione uma frota"); return; }')
    p.append('  var sel = document.getElementById("fleet-select");')
    p.append('  var fname = "relatorio_" + sel.options[sel.selectedIndex].text + ".pdf";')
    p.append('  var opt = {')
    p.append('    margin: 10,')
    p.append('    filename: fname,')
    p.append('    image: { type: "jpeg", quality: 0.98 },')
    p.append('    html2canvas: { scale: 2 },')
    p.append('    jsPDF: { orientation: "landscape", unit: "mm", format: "a4" }')
    p.append('  };')
    p.append('  html2pdf().set(opt).from(active).save();')
    p.append('}')
    p.append('</script>')

    p.append('</body>')
    p.append('</html>')

    return ''.join(p)


def generate_fleet_tab(fleet_data, tab_id, fleet_name, df_meta):
    """Generate content for a fleet dashboard - split by month"""
    p = []

    failures = fleet_data['failures']
    eq_failures = fleet_data['eq_failures']
    sys_failures = fleet_data['sys_failures']
    eq_hours = fleet_data['eq_hours']
    eq_avg = fleet_data['eq_avg']
    fleet_daily_df = fleet_data.get('fleet_daily_df', {})
    acumulado = fleet_data.get('acumulado', None)

    if not failures:
        p.append('<div id="fleet_{}" class="fleet-content active">'.format(tab_id))
        p.append('<p>Sem dados disponíveis para esta frota.</p>')
        p.append('</div>')
        return ''.join(p)

    active_class = ' active' if tab_id == 'moto' else ''
    p.append('<div id="fleet_{}" class="fleet-content{}">'.format(tab_id, active_class))

    MONTH_NAMES = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Marco', 4: 'Abril', 5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'}

    all_dates = [f['data'] for f in failures]
    if not all_dates:
        p.append('</div>')
        return ''.join(p)

    last_date = max(all_dates)
    current_month = last_date.month
    current_year = last_date.year

    if current_month == 1:
        prev_month, prev_year = 12, current_year - 1
    else:
        prev_month, prev_year = current_month - 1, current_year

    failures_prev = [f for f in failures if f['data'].month == prev_month and f['data'].year == prev_year]
    failures_curr = [f for f in failures if f['data'].month == current_month and f['data'].year == current_year]

    df_dates_prev = {k: v for k, v in fleet_daily_df.items() if k[5:7] == str(prev_month).zfill(2)}
    df_dates_curr = {k: v for k, v in fleet_daily_df.items() if k[5:7] == str(current_month).zfill(2)}

    avg_fleet_df_prev = 0
    if df_dates_prev:
        avg_fleet_df_prev = sum(df_dates_prev.values()) / len(df_dates_prev)
    elif acumulado is not None:
        avg_fleet_df_prev = acumulado

    last_7_fleet_vals = []
    sorted_fleet_dates = sorted(fleet_daily_df.keys())
    if len(sorted_fleet_dates) >= 7:
        for i in range(-7, 0):
            last_7_fleet_vals.append(fleet_daily_df[sorted_fleet_dates[i]])
    elif sorted_fleet_dates:
        last_7_fleet_vals = [fleet_daily_df[k] for k in sorted_fleet_dates]

    last_7_days_df = 0
    if last_7_fleet_vals:
        last_7_days_df = sum(last_7_fleet_vals) / len(last_7_fleet_vals)

    failures_last_7_days = 0
    if sorted_fleet_dates:
        last_7_dates_set = set(sorted_fleet_dates[-7:] if len(sorted_fleet_dates) >= 7 else sorted_fleet_dates)
        for f in failures:
            if f['data'].strftime('%Y-%m-%d') in last_7_dates_set:
                failures_last_7_days += 1

    total_failures_prev = len(failures_prev)

    df_above_meta_prev = 'ACIMA' if avg_fleet_df_prev >= df_meta else 'ABAIXO'
    df_color_prev = '#27ae60' if avg_fleet_df_prev >= df_meta else '#e74c3c'
    status_class_prev = 'above' if avg_fleet_df_prev >= df_meta else 'below'

    if last_7_days_df > 95:
        trend_status = 'EXCELENTE'
        trend_class = 'excellent'
    elif last_7_days_df > 90:
        trend_status = 'BOM'
        trend_class = 'good'
    elif last_7_days_df > 80:
        trend_status = 'ATENCAO'
        trend_class = 'attention'
    else:
        trend_status = 'CRITICO'
        trend_class = 'critical'

    p.append('<div class="section">')
    p.append('<h2 style="color: #e67e22; border-bottom-color: #e67e22;">&#9889; Resumo Executivo</h2>')
    p.append('<div class="kpi-row">')

    p.append('<div class="kpi-card">')
    p.append('<div class="kpi-label">DF ACUMULADA (MES)</div>')
    p.append('<div class="kpi-value">{:.2f}<span class="kpi-unit">%</span></div>'.format(avg_fleet_df_prev))
    p.append('<span class="kpi-status {}">{} da meta</span>'.format(status_class_prev, df_above_meta_prev))
    p.append('</div>')

    p.append('<div class="kpi-card">')
    p.append('<div class="kpi-label">DF ULTIMOS 7 DIAS</div>')
    p.append('<div class="kpi-value">{:.2f}<span class="kpi-unit">%</span></div>'.format(last_7_days_df))
    p.append('<span class="kpi-status {}">{}</span>'.format(trend_class, trend_status))
    p.append('</div>')

    p.append('<div class="kpi-card">')
    p.append('<div class="kpi-label">TOTAL DE FALHAS</div>')
    p.append('<div class="kpi-value">{}</div>'.format(total_failures_prev))
    p.append('</div>')

    p.append('<div class="kpi-card">')
    p.append('<div class="kpi-label">FALHAS (7 DIAS)</div>')
    p.append('<div class="kpi-value">{}</div>'.format(failures_last_7_days))
    p.append('</div>')

    p.append('</div>')

    critical_patterns = get_critical_failures(failures_prev, eq_hours, 3)
    num_critical = len([p2 for p2 in critical_patterns if p2[2] > 3])
    if avg_fleet_df_prev >= df_meta:
        if num_critical > 0:
            status_text = 'BOAS NOTICIAS: A frota esta <strong>acima da meta</strong> de DF ({:.2f}% vs {:.2f}%). Porem, identificamos <strong>{} padroes de falhas criticas</strong> que precisam acao imediata para evitar quedas na proxima semana.'.format(avg_fleet_df_prev, df_meta, num_critical)
        else:
            status_text = 'BOAS NOTICIAS: A frota esta <strong>acima da meta</strong> de DF ({:.2f}% vs {:.2f}%). Desempenho dentro dos parametros esperados.'.format(avg_fleet_df_prev, df_meta)
    else:
        status_text = 'ATENCAO: A frota esta <strong>abaixo da meta</strong> de DF ({:.2f}% vs {:.2f}%). Recomenda-se intensificar manutencao preventiva e investigar causas das falhas criticas.'.format(avg_fleet_df_prev, df_meta)
    p.append('<div class="status-comment">')
    p.append('<strong>Status Geral:</strong> ' + status_text)
    p.append('</div>')
    p.append('</div>')

    p.append('<div class="section">')
    p.append('<h2>Disponibilidade Fisica (DF)</h2>')

    p.append('<div class="subsection">')
    p.append('<h3>Performance do Mes</h3>')

    min_daily_avg = 100
    min_date = None
    max_daily_avg = 0
    max_date = None
    days_above = 0
    days_below = 0

    for date_key, df_val in df_dates_prev.items():
        if df_val > max_daily_avg:
            max_daily_avg = df_val
            max_date = date_key
        if df_val < min_daily_avg:
            min_daily_avg = df_val
            min_date = date_key
        if df_val >= df_meta:
            days_above += 1
        else:
            days_below += 1

    total_days = days_above + days_below

    p.append('<div class="metric-item"><span class="metric-label">DF Acumulada:</span><span class="metric-value">{:.2f}%</span></div>'.format(avg_fleet_df_prev))
    p.append('<div class="metric-item"><span class="metric-label">Comparacao a meta:</span><span class="metric-value">{:.2f}% {} da meta</span></div>'.format(abs(avg_fleet_df_prev - df_meta), df_above_meta_prev))

    max_date_str = datetime.strptime(max_date, '%Y-%m-%d').strftime('%d/%m') if max_date else 'N/A'
    p.append('<div class="metric-item"><span class="metric-label">DF Maxima:</span><span class="metric-value">{:.2f}% ({}) </span></div>'.format(max_daily_avg, max_date_str))

    min_date_str = datetime.strptime(min_date, '%Y-%m-%d').strftime('%d/%m') if min_date else 'N/A'
    p.append('<div class="metric-item"><span class="metric-label">DF Minima:</span><span class="metric-value">{:.2f}% ({}) - Dia critico</span></div>'.format(min_daily_avg, min_date_str))

    pct_above = round(100 * days_above / total_days, 1) if total_days > 0 else 0
    pct_below = round(100 * days_below / total_days, 1) if total_days > 0 else 0
    p.append('<div class="metric-item"><span class="metric-label">Dias acima da meta:</span><span class="metric-value">{} de {} ({:.1f}%)</span></div>'.format(days_above, total_days, pct_above))
    p.append('<div class="metric-item"><span class="metric-label">Dias abaixo da meta:</span><span class="metric-value">{} dias ({:.1f}%)</span></div>'.format(days_below, pct_below))

    p.append('</div>')

    p.append('<div class="subsection">')
    p.append('<h3>Ultimos 7 Dias (Tendencia)</h3>')

    days_above_7 = sum(1 for val in last_7_fleet_vals if val >= df_meta)

    if len(last_7_fleet_vals) >= 2:
        first_5_days = sum(last_7_fleet_vals[:5]) / 5 if len(last_7_fleet_vals) >= 5 else sum(last_7_fleet_vals[:2]) / 2
        last_2_days = sum(last_7_fleet_vals[-2:]) / 2
        trend_change = last_2_days - first_5_days
        if trend_change > 2:
            trend_text = 'MELHORA CONSISTENTE'
        elif trend_change < -2:
            trend_text = 'PIORA CONSISTENTE'
        else:
            trend_text = 'ESTAVEL'
    else:
        trend_text = 'DADOS INSUFICIENTES'
        trend_change = 0

    p.append('<div class="metric-item"><span class="metric-label">DF Acumulada ultimos 7 dias:</span><span class="metric-value">{:.2f}%</span></div>'.format(last_7_days_df))
    p.append('<div class="metric-item"><span class="metric-label">Tendencia:</span><span class="metric-value">{} ({:+.1f}%)</span></div>'.format(trend_text, trend_change))
    p.append('<div class="metric-item"><span class="metric-label">Dias acima da meta:</span><span class="metric-value">{} de 7</span></div>'.format(days_above_7))

    p.append('</div>')

    critical_days = []
    for date_key, df_val in df_dates_prev.items():
        if df_val < 85:
            critical_days.append((date_key, df_val))

    if critical_days:
        p.append('<div class="subsection">')
        p.append('<h3>Dias Criticos (DF &lt; 85%)</h3>')
        p.append('<div class="days-critical">')

        for date_key, df_val in sorted(critical_days)[-10:]:
            date_obj = datetime.strptime(date_key, '%Y-%m-%d')
            date_display = date_obj.strftime('%d/%m')

            longest_fail = None
            longest_duration = 0
            for f in failures_prev:
                if f['data'] == date_obj.date() and f['duracao'] > longest_duration:
                    longest_duration = f['duracao']
                    longest_fail = f

            failure_text = ''
            if longest_fail:
                failure_text = 'Falha: ' + longest_fail['system']
                if longest_fail['subsystem']:
                    failure_text += ' - ' + longest_fail['subsystem']

            p.append('<div class="day-card">')
            p.append('<div class="day-card-date">{}</div>'.format(date_display))
            p.append('<div class="day-card-df">DF: {:.2f}%</div>'.format(df_val))
            if failure_text:
                p.append('<div class="day-card-failure">{}</div>'.format(failure_text))
            p.append('</div>')

        p.append('</div>')
        p.append('</div>')

    p.append('<div class="subsection">')
    p.append('<h3>Grafico DF Diario</h3>')
    p.append('<div class="chart-container">')
    p.append('<canvas id="df_{}"></canvas>'.format(tab_id))
    p.append('</div>')
    p.append('<script>')
    p.append('var ctx_df_{} = document.getElementById("df_{}").getContext("2d");'.format(tab_id, tab_id))

    sorted_fleet_dates_all = sorted(fleet_daily_df.keys())
    df_chart_values = [round(fleet_daily_df[k], 2) for k in sorted_fleet_dates_all]

    dates_json = _json.dumps(sorted_fleet_dates_all)
    df_json = _json.dumps(df_chart_values)
    meta_json = _json.dumps([df_meta] * len(sorted_fleet_dates_all))

    p.append('new Chart(ctx_df_{}, {{'.format(tab_id))
    p.append('  type: "line",')
    p.append('  data: {')
    p.append('    labels: ' + dates_json + ',')
    p.append('    datasets: [')
    p.append('      {')
    p.append('        label: "DF Diario Frota",')
    p.append('        data: ' + df_json + ',')
    p.append('        borderColor: "#27ae60",')
    p.append('        borderWidth: 2,')
    p.append('        fill: false,')
    p.append('        tension: 0.3')
    p.append('      },')
    p.append('      {')
    p.append('        label: "Meta (' + str(df_meta) + '%)",')
    p.append('        data: ' + meta_json + ',')
    p.append('        borderColor: "#e74c3c",')
    p.append('        borderDash: [5, 5],')
    p.append('        borderWidth: 1,')
    p.append('        fill: false')
    p.append('      }')
    p.append('    ]')
    p.append('  },')
    p.append('  options: {')
    p.append('    responsive: true,')
    p.append('    maintainAspectRatio: false,')
    p.append('    scales: { y: { beginAtZero: true, max: 100 } },')
    p.append('    plugins: { legend: { display: true } }')
    p.append('  }')
    p.append('});')
    p.append('</script>')
    p.append('</div>')

    p.append('</div>')

    p.append('<div class="section">')
    p.append('<h2>Falhas</h2>')

    prefix = 'MM' if tab_id == 'moto' else 'EM'
    prev_month_name = MONTH_NAMES.get(prev_month, '')

    p.append('<div class="subsection">')
    p.append('<h3>' + prev_month_name + ' - Corretivas por Equipamento</h3>')
    p.append('<div class="eq-grid">')

    prev_eq_failures = defaultdict(int)
    prev_eq_hours = defaultdict(int)
    for f in failures_prev:
        prev_eq_failures[f['eq_id']] += 1
        prev_eq_hours[f['eq_id']] += f['duracao']

    sorted_eq = sorted(prev_eq_failures.items(), key=lambda x: x[1], reverse=True)
    if not sorted_eq:
        sorted_eq = sorted(eq_avg.items(), key=lambda x: x[0])

    for eq_id, fail_count in sorted_eq:
        df_pct = eq_avg.get(eq_id, 0)
        hours = prev_eq_hours.get(eq_id, 0) / 3600.0

        if df_pct >= 90:
            color = 'green'
        elif df_pct >= 50:
            color = 'orange'
        elif df_pct > 0:
            color = 'red'
        else:
            color = 'blue'

        eq_display = prefix + eq_id

        p.append('<div class="eq-card {0}">'.format(color))
        p.append('<div class="eq-name {0}">{1}</div>'.format(color, eq_display))
        p.append('<div class="eq-stat"><span class="eq-stat-label">Corretivas:</span><span class="eq-stat-value">{}</span></div>'.format(fail_count))
        p.append('<div class="eq-stat"><span class="eq-stat-label">Horas Paradas:</span><span class="eq-stat-value">{}</span></div>'.format(int(hours)))
        p.append('<div class="eq-stat"><span class="eq-stat-label">DF %:</span><span class="eq-stat-value">{:.1f}</span></div>'.format(df_pct))
        p.append('<div class="progress-bar"><div class="progress-fill" style="width: {}%;"></div></div>'.format(min(df_pct, 100)))
        p.append('</div>')

    p.append('</div>')
    p.append('</div>')

    prev_sys_failures = defaultdict(int)
    for f in failures_prev:
        if f['system']:
            prev_sys_failures[f['system']] += 1

    p.append('<div class="subsection">')
    p.append('<h3>' + prev_month_name + ' - Pareto de Falhas - Frequencia</h3>')
    p.append('<div class="chart-container">')
    p.append('<canvas id="pareto_{}_prev"></canvas>'.format(tab_id))
    p.append('</div>')
    p.append('<script>')
    p.append('var ctx_{}_prev = document.getElementById("pareto_{}_prev").getContext("2d");'.format(tab_id, tab_id))

    pareto_prev, _ = get_pareto_data(prev_sys_failures)
    pareto_labels = [item[0] for item in pareto_prev]
    pareto_data = [item[1] for item in pareto_prev]

    cum_data = []
    cum = 0
    total = sum(pareto_data)
    for val in pareto_data:
        cum += val
        cum_pct = int(100 * cum / total) if total > 0 else 0
        cum_data.append(cum_pct)

    labels_json = _json.dumps(pareto_labels)
    data_json = _json.dumps(pareto_data)
    cum_json = _json.dumps(cum_data)

    p.append('new Chart(ctx_{}_prev, {{'.format(tab_id))
    p.append('  type: "bar",')
    p.append('  data: {')
    p.append('    labels: ' + labels_json + ',')
    p.append('    datasets: [')
    p.append('      {')
    p.append('        label: "Frequencia",')
    p.append('        data: ' + data_json + ',')
    p.append('        backgroundColor: "#1a3a52",')
    p.append('        order: 2')
    p.append('      },')
    p.append('      {')
    p.append('        label: "Acumulado %",')
    p.append('        data: ' + cum_json + ',')
    p.append('        type: "line",')
    p.append('        borderColor: "#e74c3c",')
    p.append('        borderWidth: 2,')
    p.append('        fill: false,')
    p.append('        yAxisID: "y1",')
    p.append('        order: 1')
    p.append('      }')
    p.append('    ]')
    p.append('  },')
    p.append('  options: {')
    p.append('    responsive: true,')
    p.append('    maintainAspectRatio: false,')
    p.append('    scales: {')
    p.append('      y: { beginAtZero: true, title: { display: true, text: "Frequencia" } },')
    p.append('      y1: { type: "linear", position: "right", max: 100, title: { display: true, text: "Acumulado %" } }')
    p.append('    },')
    p.append('    plugins: { legend: { display: true }, title: { display: false } }')
    p.append('  }')
    p.append('});')
    p.append('</script>')
    p.append('</div>')

    p.append('<div class="subsection">')
    p.append('<h3>' + prev_month_name + ' - Pareto por Horas Paradas</h3>')
    p.append('<div class="chart-container">')
    p.append('<canvas id="hours_{}_prev"></canvas>'.format(tab_id))
    p.append('</div>')
    p.append('<script>')
    p.append('var ctx_hours_{}_prev = document.getElementById("hours_{}_prev").getContext("2d");'.format(tab_id, tab_id))

    eq_hours_sorted = sorted(prev_eq_hours.items(), key=lambda x: x[1], reverse=True)
    hours_labels = [prefix + eq for eq, _ in eq_hours_sorted]
    hours_values = [h / 3600.0 for _, h in eq_hours_sorted]

    hours_labels_json = _json.dumps(hours_labels)
    hours_json = _json.dumps([round(h, 1) for h in hours_values])

    p.append('new Chart(ctx_hours_{}_prev, {{'.format(tab_id))
    p.append('  type: "bar",')
    p.append('  data: {')
    p.append('    labels: ' + hours_labels_json + ',')
    p.append('    datasets: [{')
    p.append('      label: "Horas Paradas",')
    p.append('      data: ' + hours_json + ',')
    p.append('      backgroundColor: "#e74c3c"')
    p.append('    }]')
    p.append('  },')
    p.append('  options: {')
    p.append('    indexAxis: "x",')
    p.append('    responsive: true,')
    p.append('    maintainAspectRatio: false,')
    p.append('    scales: { y: { beginAtZero: true } },')
    p.append('    plugins: { legend: { display: true } }')
    p.append('  }')
    p.append('});')
    p.append('</script>')
    p.append('</div>')

    p.append('</div>')

    critical_prev = get_critical_failures(failures_prev, prev_eq_hours, 3)
    if critical_prev:
        p.append('<div class="subsection">')
        p.append('<h3>' + prev_month_name + ' - Falhas Criticas Identificadas</h3>')

        for i, (pattern_name, count, hours, dates) in enumerate(critical_prev):
            severity = 'ALTA' if hours > 10 else ('MEDIA' if hours > 3 else 'BAIXA')
            sev_color = '#e74c3c' if severity == 'ALTA' else ('#f39c12' if severity == 'MEDIA' else '#3498db')
            eq_affected = set()
            for f in failures_prev:
                fkey = f['system'] + ' - ' + f['subsystem'] if f['subsystem'] else f['system']
                if fkey == pattern_name:
                    eq_affected.add(prefix + f['eq_id'])
            p.append('<div class="critical-card" style="border-left-color: {};">'.format(sev_color))
            p.append('<h3 style="color: {};">{}. {} <span class="action-badge" style="background: {}; color: white;">{}</span></h3>'.format(sev_color, i + 1, pattern_name, sev_color, severity))
            p.append('<p><strong>Ocorrencias:</strong> {} eventos | <strong>Tempo Total Parado:</strong> {:.1f}h</p>'.format(count, hours))
            p.append('<p><strong>Equipamentos Afetados:</strong> {}</p>'.format(', '.join(sorted(eq_affected))))
            p.append('<p><strong>Datas:</strong> {}</p>'.format(', '.join(dates[:8])))
            if hours > 5:
                p.append('<p style="color: #e74c3c; margin-top: 8px;"><strong>Impacto:</strong> Perda significativa de disponibilidade. Requer plano de acao imediato com analise de causa raiz.</p>')
            elif count >= 3:
                p.append('<p style="color: #f39c12; margin-top: 8px;"><strong>Impacto:</strong> Padrao recorrente identificado. Avaliar intervalo de manutencao preventiva.</p>')
            p.append('</div>')

        p.append('</div>')

    p.append('<div class="subsection">')
    p.append('<h3>' + prev_month_name + ' - Padroes Recorrentes</h3>')
    pattern_groups_prev = defaultdict(lambda: {'count': 0, 'eqs': set(), 'hours': 0})
    for f in failures_prev:
        if f['system']:
            pattern_groups_prev[f['system']]['count'] += 1
            pattern_groups_prev[f['system']]['eqs'].add(f['eq_id'])
            pattern_groups_prev[f['system']]['hours'] += f['duracao'] / 3600.0
    recurrent_prev = [(sys, d) for sys, d in pattern_groups_prev.items() if d['count'] >= 3 or len(d['eqs']) >= 2]
    recurrent_prev.sort(key=lambda x: x[1]['count'], reverse=True)
    if recurrent_prev:
        for sys_name, d in recurrent_prev[:5]:
            eq_list = ', '.join([prefix + e for e in sorted(d['eqs'])])
            classification = 'RECORRENTE' if d['count'] >= 5 else ('POTENCIAL' if d['count'] >= 3 else 'PONTUAL')
            cls_color = '#e74c3c' if classification == 'RECORRENTE' else ('#f39c12' if classification == 'POTENCIAL' else '#3498db')
            p.append('<div style="padding: 12px; margin-bottom: 10px; border-left: 3px solid {}; background: #fafafa; border-radius: 4px;">'.format(cls_color))
            p.append('<strong style="color: {};">{}</strong> - <span class="action-badge" style="background: {}; color: white;">{}</span>'.format(cls_color, sys_name, cls_color, classification))
            p.append('<p style="font-size: 13px; margin-top: 6px;">{} ocorrencias | {:.1f}h paradas | Equipamentos: {}</p>'.format(d['count'], d['hours'], eq_list))
            p.append('</div>')
    else:
        p.append('<p>Nenhum padrao recorrente identificado.</p>')
    p.append('</div>')

    if failures_curr:
        curr_month_name = MONTH_NAMES.get(current_month, '')
        today = datetime.now().date()
        curr_date_str = '01/{} a {}/{}'.format(str(current_month).zfill(2), str(today.day).zfill(2), str(current_month).zfill(2))

        p.append('<div class="subsection">')
        p.append('<h3>Analise Parcial - ' + curr_month_name + ' (' + curr_date_str + ')</h3>')
        p.append('<div class="eq-grid">')

        curr_eq_failures = defaultdict(int)
        curr_eq_hours = defaultdict(int)
        for f in failures_curr:
            curr_eq_failures[f['eq_id']] += 1
            curr_eq_hours[f['eq_id']] += f['duracao']

        sorted_eq_curr = sorted(curr_eq_failures.items(), key=lambda x: x[1], reverse=True)
        for eq_id, fail_count in sorted_eq_curr:
            df_pct = eq_avg.get(eq_id, 0)
            hours = curr_eq_hours.get(eq_id, 0) / 3600.0

            if df_pct >= 90:
                color = 'green'
            elif df_pct >= 50:
                color = 'orange'
            elif df_pct > 0:
                color = 'red'
            else:
                color = 'blue'

            eq_display = prefix + eq_id

            p.append('<div class="eq-card {0}">'.format(color))
            p.append('<div class="eq-name {0}">{1}</div>'.format(color, eq_display))
            p.append('<div class="eq-stat"><span class="eq-stat-label">Corretivas:</span><span class="eq-stat-value">{}</span></div>'.format(fail_count))
            p.append('<div class="eq-stat"><span class="eq-stat-label">Horas Paradas:</span><span class="eq-stat-value">{}</span></div>'.format(int(hours)))
            p.append('<div class="eq-stat"><span class="eq-stat-label">DF %:</span><span class="eq-stat-value">{:.1f}</span></div>'.format(df_pct))
            p.append('<div class="progress-bar"><div class="progress-fill" style="width: {}%;"></div></div>'.format(min(df_pct, 100)))
            p.append('</div>')

        p.append('</div>')
        p.append('</div>')

        curr_sys_failures = defaultdict(int)
        for f in failures_curr:
            if f['system']:
                curr_sys_failures[f['system']] += 1

        p.append('<div class="subsection">')
        p.append('<h3>' + curr_month_name + ' - Pareto de Falhas - Frequencia</h3>')
        p.append('<div class="chart-container">')
        p.append('<canvas id="pareto_{}_curr"></canvas>'.format(tab_id))
        p.append('</div>')
        p.append('<script>')
        p.append('var ctx_{}_curr = document.getElementById("pareto_{}_curr").getContext("2d");'.format(tab_id, tab_id))

        pareto_curr, _ = get_pareto_data(curr_sys_failures)
        pareto_labels_curr = [item[0] for item in pareto_curr]
        pareto_data_curr = [item[1] for item in pareto_curr]

        cum_data_curr = []
        cum_curr = 0
        total_curr = sum(pareto_data_curr)
        for val in pareto_data_curr:
            cum_curr += val
            cum_pct_curr = int(100 * cum_curr / total_curr) if total_curr > 0 else 0
            cum_data_curr.append(cum_pct_curr)

        labels_json_curr = _json.dumps(pareto_labels_curr)
        data_json_curr = _json.dumps(pareto_data_curr)
        cum_json_curr = _json.dumps(cum_data_curr)

        p.append('new Chart(ctx_{}_curr, {{'.format(tab_id))
        p.append('  type: "bar",')
        p.append('  data: {')
        p.append('    labels: ' + labels_json_curr + ',')
        p.append('    datasets: [')
        p.append('      {')
        p.append('        label: "Frequencia",')
        p.append('        data: ' + data_json_curr + ',')
        p.append('        backgroundColor: "#1a3a52",')
        p.append('        order: 2')
        p.append('      },')
        p.append('      {')
        p.append('        label: "Acumulado %",')
        p.append('        data: ' + cum_json_curr + ',')
        p.append('        type: "line",')
        p.append('        borderColor: "#e74c3c",')
        p.append('        borderWidth: 2,')
        p.append('        fill: false,')
        p.append('        yAxisID: "y1",')
        p.append('        order: 1')
        p.append('      }')
        p.append('    ]')
        p.append('  },')
        p.append('  options: {')
        p.append('    responsive: true,')
        p.append('    maintainAspectRatio: false,')
        p.append('    scales: {')
        p.append('      y: { beginAtZero: true, title: { display: true, text: "Frequencia" } },')
        p.append('      y1: { type: "linear", position: "right", max: 100, title: { display: true, text: "Acumulado %" } }')
        p.append('    },')
        p.append('    plugins: { legend: { display: true }, title: { display: false } }')
        p.append('  }')
        p.append('});')
        p.append('</script>')
        p.append('</div>')

        p.append('<div class="subsection">')
        p.append('<h3>' + curr_month_name + ' - Pareto por Horas Paradas</h3>')
        p.append('<div class="chart-container">')
        p.append('<canvas id="hours_{}_curr"></canvas>'.format(tab_id))
        p.append('</div>')
        p.append('<script>')
        p.append('var ctx_hours_{}_curr = document.getElementById("hours_{}_curr").getContext("2d");'.format(tab_id, tab_id))

        eq_hours_sorted_curr = sorted(curr_eq_hours.items(), key=lambda x: x[1], reverse=True)
        hours_labels_curr = [prefix + eq for eq, _ in eq_hours_sorted_curr]
        hours_values_curr = [h / 3600.0 for _, h in eq_hours_sorted_curr]

        hours_labels_json_curr = _json.dumps(hours_labels_curr)
        hours_json_curr = _json.dumps([round(h, 1) for h in hours_values_curr])

        p.append('new Chart(ctx_hours_{}_curr, {{'.format(tab_id))
        p.append('  type: "bar",')
        p.append('  data: {')
        p.append('    labels: ' + hours_labels_json_curr + ',')
        p.append('    datasets: [{')
        p.append('      label: "Horas Paradas",')
        p.append('      data: ' + hours_json_curr + ',')
        p.append('      backgroundColor: "#e74c3c"')
        p.append('    }]')
        p.append('  },')
        p.append('  options: {')
        p.append('    indexAxis: "x",')
        p.append('    responsive: true,')
        p.append('    maintainAspectRatio: false,')
        p.append('    scales: { y: { beginAtZero: true } },')
        p.append('    plugins: { legend: { display: true } }')
        p.append('  }')
        p.append('});')
        p.append('</script>')
        p.append('</div>')

    p.append('<div class="section">')
    p.append('<h2>Conclusao - {}</h2>'.format(fleet_name))
    p.append('<div style="padding: 15px; background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); border-radius: 6px;">')
    p.append('<p style="font-size: 14px; line-height: 1.8;">')
    p.append('A frota de <strong>{}</strong> registrou <strong>{} falhas corretivas</strong> em {}, '.format(fleet_name, total_failures_prev, prev_month_name))
    total_hours_prev = sum(f['duracao'] for f in failures_prev) / 3600.0
    p.append('totalizando <strong>{:.1f} horas paradas</strong>. '.format(total_hours_prev))
    p.append('O DF acumulado da frota foi de <strong style="color: {};">{:.2f}%</strong> (<strong style="color: {};">{}</strong>, meta: {}%). '.format(df_color_prev, avg_fleet_df_prev, df_color_prev, df_above_meta_prev, df_meta))
    if critical_prev:
        p.append('As principais falhas criticas foram no sistema <strong>{}</strong> com {:.1f}h de parada. '.format(critical_prev[0][0], critical_prev[0][2]))
    p.append('</p>')
    p.append('<p style="font-size: 12px; color: #888; margin-top: 10px;">Engenharia de Manutencao e Confiabilidade - Setor GRD</p>')
    p.append('</div>')
    p.append('</div>')

    p.append('</div>')  # closes fleet-content div

    return ''.join(p)


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    """Render initial upload page"""
    return '''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Upload - Relatorio de Confiabilidade</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { font-family: Segoe UI, Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #1a3a52 0%, #2d5a7a 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
            .upload-container { background: white; padding: 40px; border-radius: 8px; box-shadow: 0 8px 32px rgba(0,0,0,0.2); max-width: 600px; width: 100%; }
            .upload-container h1 { color: #1a3a52; margin-bottom: 10px; font-size: 28px; }
            .upload-container p { color: #666; margin-bottom: 30px; font-size: 14px; line-height: 1.6; }
            .form-group { margin-bottom: 25px; }
            .form-group label { display: block; margin-bottom: 10px; color: #333; font-weight: 500; }
            .file-input-wrapper { position: relative; overflow: hidden; display: inline-block; width: 100%; }
            .file-input-wrapper input[type=file] { position: absolute; left: -9999px; }
            .file-input-label { display: block; padding: 20px; border: 2px dashed #1a3a52; border-radius: 6px; text-align: center; cursor: pointer; transition: all 0.3s; background: #f9f9f9; }
            .file-input-label:hover { background: #f0f0f0; border-color: #2d5a7a; }
            .file-input-label.active { background: #e8f0ff; border-color: #1a3a52; }
            .file-name { margin-top: 10px; color: #27ae60; font-size: 14px; }
            .meta-section { background: #f0f4f8; border: 1px solid #d0d8e0; border-radius: 6px; padding: 20px; margin-bottom: 25px; }
            .meta-section h3 { color: #1a3a52; font-size: 16px; margin-bottom: 15px; }
            .meta-row { display: flex; gap: 20px; }
            .meta-field { flex: 1; }
            .meta-field label { display: block; font-size: 13px; color: #555; margin-bottom: 6px; font-weight: 500; }
            .meta-field input { width: 100%; padding: 10px 12px; border: 1px solid #ccc; border-radius: 4px; font-size: 14px; }
            .meta-field input:focus { outline: none; border-color: #1a3a52; box-shadow: 0 0 0 2px rgba(26,58,82,0.15); }
            .meta-field .hint { font-size: 11px; color: #888; margin-top: 4px; }
            .submit-btn { background: linear-gradient(135deg, #1a3a52 0%, #2d5a7a 100%); color: white; padding: 14px 32px; border: none; border-radius: 6px; cursor: pointer; font-size: 16px; font-weight: 600; width: 100%; transition: transform 0.2s; }
            .submit-btn:hover { transform: translateY(-2px); box-shadow: 0 6px 16px rgba(26, 58, 82, 0.3); }
            .submit-btn:disabled { opacity: 0.5; cursor: not-allowed; transform: none; }
            .loading { display: none; text-align: center; color: #1a3a52; margin-top: 20px; }
            .spinner { border: 4px solid #f3f3f3; border-top: 4px solid #1a3a52; border-radius: 50%; width: 40px; height: 40px; animation: spin 1s linear infinite; margin: 0 auto 10px; }
            @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
        </style>
    </head>
    <body>
        <div class="upload-container">
            <h1>Relatorio de Confiabilidade</h1>
            <p>Analyze and generate reliability reports from failure data Excel files. Upload your data file to get started.</p>
            <form id="uploadForm" enctype="multipart/form-data">
                <div class="form-group">
                    <label>Selecione arquivo Excel</label>
                    <div class="file-input-wrapper">
                        <input type="file" id="fileInput" name="file" accept=".xlsx,.xls" required>
                        <label for="fileInput" class="file-input-label" id="fileLabel">
                            <strong>Clique para selecionar</strong><br>ou arraste o arquivo aqui
                        </label>
                        <div class="file-name" id="fileName"></div>
                    </div>
                </div>
                <div class="meta-section">
                    <h3>Meta de DF por Frota</h3>
                    <div class="meta-row">
                        <div class="meta-field">
                            <label>Motoniveladora (%)</label>
                            <input type="number" id="metaMoto" name="meta_moto" value="72.00" step="0.01" min="0" max="100">
                            <div class="hint">Padrao: 72,00%</div>
                        </div>
                        <div class="meta-field">
                            <label>Escavadeira (%)</label>
                            <input type="number" id="metaEsc" name="meta_esc" value="90.50" step="0.01" min="0" max="100">
                            <div class="hint">Padrao: 90,50%</div>
                        </div>
                    </div>
                </div>
                <button type="submit" class="submit-btn" id="submitBtn">Processar Relatorio</button>
                <div class="loading" id="loading">
                    <div class="spinner"></div>
                    <p>Processando dados...</p>
                </div>
            </form>
        </div>
        <script>
            var fileInput = document.getElementById("fileInput");
            var fileLabel = document.getElementById("fileLabel");
            var fileName = document.getElementById("fileName");

            fileInput.addEventListener("change", function() {
                if (this.files.length > 0) {
                    fileName.textContent = "Arquivo: " + this.files[0].name;
                    fileLabel.classList.add("active");
                }
            });

            fileLabel.addEventListener("dragover", function(e) {
                e.preventDefault();
                this.classList.add("active");
            });

            fileLabel.addEventListener("dragleave", function(e) {
                e.preventDefault();
                if (!fileInput.files.length) {
                    this.classList.remove("active");
                }
            });

            fileLabel.addEventListener("drop", function(e) {
                e.preventDefault();
                fileInput.files = e.dataTransfer.files;
                fileInput.dispatchEvent(new Event("change"));
            });

            document.getElementById("uploadForm").addEventListener("submit", function(e) {
                e.preventDefault();
                if (!fileInput.files.length) {
                    alert("Selecione um arquivo");
                    return;
                }

                var formData = new FormData();
                formData.append("file", fileInput.files[0]);
                formData.append("meta_moto", document.getElementById("metaMoto").value);
                formData.append("meta_esc", document.getElementById("metaEsc").value);

                document.getElementById("submitBtn").disabled = true;
                document.getElementById("loading").style.display = "block";

                fetch("/api/upload", {
                    method: "POST",
                    body: formData
                })
                .then(r => r.json())
                .then(data => {
                    if (data.success) {
                        window.location.href = "/report?id=" + data.report_id;
                    } else {
                        alert("Erro: " + data.error);
                        document.getElementById("submitBtn").disabled = false;
                        document.getElementById("loading").style.display = "none";
                    }
                })
                .catch(err => {
                    alert("Erro ao processar: " + err.message);
                    document.getElementById("submitBtn").disabled = false;
                    document.getElementById("loading").style.display = "none";
                });
            });
        </script>
    </body>
    </html>
    '''


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle file upload and process data"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Arquivo vazio'}), 400

        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Formato invalido. Use .xlsx ou .xls'}), 400

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        meta_moto = request.form.get('meta_moto', '72.0')
        meta_esc = request.form.get('meta_esc', '90.50')
        try:
            meta_moto = float(meta_moto)
        except (ValueError, TypeError):
            meta_moto = 72.0
        try:
            meta_esc = float(meta_esc)
        except (ValueError, TypeError):
            meta_esc = 90.50

        custom_meta = {'moto': meta_moto, 'esc': meta_esc}

        data = process_excel(filepath)
        html_report = generate_report(data, custom_meta)

        report_id = str(int(__import__('time').time() * 1000))
        report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'report_' + report_id + '.html')

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(html_report)

        return jsonify({'success': True, 'report_id': report_id}), 200

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/report')
def report():
    """Display generated report"""
    try:
        report_id = request.args.get('id')
        if not report_id:
            return 'Report ID not provided', 400

        report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'report_' + report_id + '.html')
        if not os.path.exists(report_path):
            return 'Report not found', 404

        with open(report_path, 'r', encoding='utf-8') as f:
            return f.read()

    except Exception as e:
        return 'Error loading report: ' + str(e), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
